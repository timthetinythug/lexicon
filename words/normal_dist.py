from openpyxl import Workbook
wb = Workbook()
wc = Workbook()

from assemb_class import Assembly

ws = wb.active
wt = wc.active

ws['A1'] = ""
ws['B1'] = "Word"
ws['C1'] = "Length"
ws['D1'] = "norm(L)"
ws['E1'] = "Frequency"
ws['F1'] = "norm(F)"

wt['A1'] = "range"
wt['B1'] = "upper Bound"
wt['C1'] = "# of words w/length == upper Bound"
wt['D1'] = "# of words w/frequency == upper Bound"


def Norm(x, mean, std_dev, bin="FALSE"):
    return "=(NORM.DIST({0}, {1}, {2}, {3})".format(x, mean, std_dev, bin)

def create_spread(series):
    books = Assembly(series)

    n = 1
    for w in books.words:
        ws.append([n, w, books.final[w][0], books.final[w][1], books.final[w][2], books.final[w][3]])
        n += 1

    temp = sorted(books.ranges.items(), key= lambda k: k[1][0])
    for i in temp:
        wt.append([i[0], i[1][0], i[1][1], i[1][2]])

    ws.append(["max", "", "=MAX(C2:C{0})".format(n), "=MAX(D2:D{0})".format(n), "=MAX(E2:E{0})".format(n),
              "=MAX(F2:F{0})".format(n)])
    ws.append(["avg", "", "=AVERAGE(C2:C{0})".format(n), "=AVERAGE(D2:D{0})".format(n),
               "=AVERAGE(E2:E{0})".format(n), "=AVERAGE(F2:F{0})".format(n)])
    ws.append(["mode", "", "=MODE(C2:C{0})".format(n), "=MODE(D2:D{0})".format(n), "=MODE(E2:E{0})".format(n),
               "=MODE(F2:F{0})".format(n)])
    ws.append(["median", "", "=MEDIAN(C2:C{0})".format(n), "=MEDIAN(D2:D{0})".format(n),
               "=MEDIAN(E2:E{0})".format(n), "=MEDIAN(F2:F{0})".format(n)])
    ws.append(["std", "", "=STDEV(C2:C{0})".format(n), "=STDEV(D2:D{0})".format(n),
               "=STDEV(E2:E{0})".format(n), "=STDEV(F2:F{0})".format(n)])

    #print(temp)

    wb.save("ex_word_distribution.xlsx")
    #wc.save("wordy.xlsx")


set_of_books = ["harry_1.txt", "harry_2.txt", "harry_3.txt", "harry_4.txt", "harry_5.txt", "harry_6.txt",
                 "harry_7.txt", "infinite_jest.txt", "da_vinci_code.txt", "grow_rich.txt"]

example = ["harry_1.txt", "harry_2.txt", "harry_3.txt"]

create_spread(example)

# Get this figure: fig = py.get_figure("https://plot.ly/~MattSundquist/19547/")
# Get this figure's data: data = py.get_figure("https://plot.ly/~MattSundquist/19547/").get_data()
# Add data to this figure: py.plot(Data([Scatter(x=[1, 2], y=[2, 3])]), filename ="histogram-mpl-legend", fileopt="extend")
# Get y data of first trace: y1 = py.get_figure("https://plot.ly/~MattSundquist/19547/").get_data()[0]["y"]

# Get figure documentation: https://plot.ly/python/get-requests/
# Add data documentation: https://plot.ly/python/file-options/

# If you're using unicode in your file, you may need to specify the encoding.
# You can reproduce this figure in Python with the following code!

# Learn about API authentication here: https://plot.ly/python/getting-started
# Find your api_key here: https://plot.ly/settings/api

#'''
#import plotly.plotly as py
#from plotly.graph_objs import *
#py.sign_in('timthetinythug', 'FUbVLyzCHZ1cP9MGeiJb')
  #
#data_x = []
#data_y1 = []
#data_y2 = []
  #
#for i in create_spread(set_of_books):
#    data_x.append(i[1][0])
#    data_y1.append(i[1][2])
#    data_y2.append(i[1][1])
  #
  #
#trace1 = {
#  "x": data_x,
#  "y": data_y1,
#  "line": {
#    "color": "#FF0000",
#    "dash": "dash",
#    "width": 1.0
#  },
#  "mode": "lines",
#  "name": "_line0",
#  "type": "scatter",
#  "xaxis": "x1",
#  "yaxis": "y1"
#}
#trace2 = {
#  "x": data_x,
#  "y": data_y2,
#  "marker": {
#    "color": "#008000",
#    "line": {"width": 1.0}
#  },
#  "opacity": 0.5,
#  "orientation": "v",
#  "type": "bar",
#  "xaxis": "x1",
#  "yaxis": "y1"
#}
#data = Data([trace1, trace2])
#layout = {
#  "bargap": 7.1054273576e-15,
#  "hovermode": "closest",
#  "showlegend": False,
#  "xaxis1": {
#    "anchor": "y1",
#    "domain": [0.0, 1.0],
#    "mirror": "ticks",
#    "nticks": 8,
#    "range": [0, 40.0],
#    "showgrid": False,
#    "showline": True,
#    "side": "bottom",
#    "tickfont": {"size": 12.0},
#    "ticks": "inside",
#    "title": "length/frequency",
#    "titlefont": {
#      "color": "#000000",
#      "size": 12.0
#    },
#    "type": "linear",
#    "zeroline": False
#  },
#  "yaxis1": {
#    "anchor": "x1",
#    "domain": [0.0, 1.0],
#    "mirror": "ticks",
#    "nticks": 7,
#    "range": [0.0, 8000.0],
#    "showgrid": False,
#    "showline": True,
#    "side": "left",
#    "tickfont": {"size": 12.0},
#    "ticks": "inside",
#    "title": "# of words",
#    "titlefont": {
#      "color": "#000000",
#      "size": 12.0
#    },
#    "type": "linear",
#    "zeroline": False
#  }
#}
#fig = Figure(data=data, layout=layout)
#plot_url = py.plot(fig)
#'''
