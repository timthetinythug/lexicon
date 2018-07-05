from isalp import notalp, art
from parser import parsed
import math
import scipy.stats as stats
import numpy as np
import sys

sys.setrecursionlimit(3232523)

class Assembly:
    def __init__(self, books, freq_moment=0, length_moment=0, concept_moment=0):
        self.books = books
        self.freq = freq_moment # weight given to freq
        self.length = length_moment # weight given to length
        self.concept = concept_moment # weight given to complexity (i.e. verb, tense, subject etc.)
        self.words = {} # {"word": freq}
        self.stats = {} # {"title": [num_new_words, avg_new_words, [list_words]]}
        self.avg_f = 0
        self.avg_l = 0
        self.dev_f = 0
        self.dev_l = 0
        self.final = {} # {"word": [std(len), std(freq)]}
        self.max_f = 0
        self.max_l = 0
        self.ranges = {} # {"from n-1 to n": [n,

        def format(txt):
            sum_length = 0 # sum of length of unique words in this book for later statistics
            prev_wordCount = len(self.words) # number of words already formatted and placed in self.words
            with open(txt) as f:
                data = f.readlines()
            new_data = list(filter(lambda k: len(k) > 3, data)) # removes white lines in b/w paragraphs
            newer_data = list(map(lambda k: k.strip(), new_data)) # removes indents and '\n' after each line
            newest_data = list(map(lambda k: k.split(), newer_data)) # example: [['the', 'cat', 'ate'], ['everyone']]
            M = []
            while newest_data != []:
                x = newest_data.pop(0)
                for i in x:
                    M.append(i.lower())
            N = list(map(lambda k: notalp(k), M)) # removes plural forms, syntax, and numbers
            O = list(filter(lambda k: type(k) != type(None), N)) # notalp('...') => None, this removes None type
            P = list(filter(lambda k: art(k), O)) # removes small words like 'i' and 'a'
            for i in P:
                if i in self.words:
                    self.words[i] += 1
                else:
                    self.words[i] = 1
                    sum_length += len(i)
            current_wordCount = len(self.words) - prev_wordCount # non important public stats
            if current_wordCount == 0:
                self.stats[str(txt)] = list([0, 0, P])
            else:
                average = sum_length / current_wordCount
                self.stats[str(txt)] = list([current_wordCount, average, P])

        for i in books:
            format(i)

        print(len(self.words))

        sum = 0
        for i in self.words:
            sum += len(i)
        self.avg_l = sum/len(self.words)

        sum = 0
        for i in self.words:
            sum += (len(i) - self.avg_l)**2
        self.dev_l = math.sqrt(sum/(len(self.words) - 1))

        sum = 0
        for i in self.words:
            sum += self.words[i]
        self.avg_f = sum/len(self.words)

        sum = 0
        for i in self.words:
            sum += (self.words[i] - self.avg_f)**2
        self.dev_f = math.sqrt(sum/(len(self.words) - 1))

        ###### FREQUENCIES HAVE HUGE OUTLIERS
        ###### THE FOLLOWING CODE IS USED TO GET RID OF THEM USING INTER-QUARTILE METHOD

        temp = sorted(self.words.items(), key= lambda k: k[1])
        out = ["fahim"]

        def IQR(x):
            z = list(map(lambda k: k[1], x))
            return np.percentile(z, 75) - np.percentile(z, 25)

        def mean(x):
            z = list(map(lambda k: k[1], x))
            return np.mean(z)

        def outliers(x, lout, T): # x must be a sorted list in ascending order
                                     # x = [(key, freq), (key, freq), ...]

            thresh = IQR(x) * T
            m = mean(x)

            first = x[0][1]
            last = x[-1][1]

            if (m - first) < (last - m):
                if thresh < (last-m):
                    return outliers(x[:-1], lout + [(x[-1][0])], T)
                else:
                    return lout
            else:
                if thresh < (m - first):
                    return outliers(x[1:], lout + [(x[0][0])], T)
                else:
                    return lout

        for i in outliers(temp, out, 15):
            if i in self.words:
                del self.words[i] # Get rid of outliers

        #########
        #########

        temp_f = sorted(self.words.items(), key=lambda k: k[1])
        self.max_f = temp_f[-1][1]

        sum = 0
        for i in self.words:
            sum += self.words[i]
        self.avg_f = sum/len(self.words)

        temp_l = list(map(lambda k: len(k), self.words.keys()))
        self.max_l = max(temp_l)

        sum = 0
        for i in self.words:
            sum += len(i)
        self.avg_l = sum/len(self.words)

        sum = 0
        for i in self.words:
            sum += (self.words[i] - self.avg_f)**2
        self.dev_f = math.sqrt(sum/(len(self.words) - 1))

        sum = 0
        for i in self.words:
            sum += (len(i) - self.avg_l)**2
        self.dev_l = math.sqrt(sum/(len(self.words) - 1))

        n = self.max_f
        range = "from {0} to {1}"

        while n != 0:
            self.ranges[range.format(n - 1, n)] = [n, 0, 0]  #[range(n-1, n), #words_w/length==n, #words_w/frequency==n)
            for i in self.words:
                k = len(i)
                if self.words[i] == n:
                    self.ranges[range.format(n-1, n)][2] += 1
                if k == n:
                    self.ranges[range.format(n-1, n)][1] += 1
            n += -1


        for i in self.words:
            self.final[i] = [len(i), stats.norm.pdf(len(i), self.avg_l, self.dev_l),
                             self.words[i], stats.norm.pdf(self.words[i], self.avg_f, self.dev_f)]
        # The norm for frequency of a word is the probability that that word is so frequent

        print(len(self.words))
        print(self.avg_l)
        print(self.avg_f)

    def __eq__(self, other):
        return self.books == other.books and \
            self.freq == other.freq and \
            self.length == other.length and \
            self.concept == other.concept

    def __repr__(self):
        sep = "------------------------------------------------------------------------------------------ \n"
        temp = "The texts included are: \n"
        for i in self.books:
            temp += str(i) + ": This book has " + str(self.stats[i][0]) + " new words " \
            + "with an average word length of " + str(self.stats[i][1])[:4] + " \n"
        return sep + temp + sep

def create_spec_words(doc):
    with open(doc) as f:
        words = f.readlines()
        data = list(map(lambda k: k[:-1], words)) # list of 3000 words --> ['art', 'Bees', ...] :: make sure lowercase
        print(data)
        print(len(data))
        return data


# def compare(lofw, wordbank) compares a list of words to a collection of words for cross-ref
# compare(lofw, wordbank): list, Assembly object --> ".txt" file
def compare(lofw, wordbank, filename):
    emp = []
    c=0
    for i in lofw:
        if i.lower() in wordbank.final:
            c += 1
            x = str(i)
            x += "\n"
            for k in wordbank.final[i.lower()]:
                x += "{0} \n".format(k)
            emp.append(x)

    with open(filename, 'w') as f:
        while len(emp) != 0:
            k = emp.pop(0)
            f.writelines(k)
            f.write("\n")

def comparev2(lofw, wordbank):
    emp = {}
    for i in lofw:
        if i.lower() in wordbank.final:
            emp[i.lower()] = wordbank.final[i.lower()]

    x = float(stats.norm.pdf(wordbank.avg_f, wordbank.avg_f, wordbank.dev_f))
    y = float(stats.norm.pdf(wordbank.avg_l, wordbank.avg_l, wordbank.dev_l))

    assignment = {}
    for i in emp:
        deviance_F = wordbank.final[i][2] - wordbank.avg_f
        if deviance_F > 0:
            if deviance_F > 3 * wordbank.dev_f:
                a = 0.80/x
                b = 0.20/y
            elif deviance_F > 2 * wordbank.dev_f:
                a = 0.70/x
                b = 0.30/y
            elif deviance_F > wordbank.dev_f:
                a = 0.60/x
                b = 0.30/y
            else:
                a = 0.45/x
                b = 0.55/y
        else:
            diff = abs(deviance_F)
            if diff > 0.5 * wordbank.avg_f:
                a = 0.35/x
                b = 0.65/y
            else:
                a = 0.45/x
                b = 0.55/y


        assignment[i] = [a * wordbank.final[i][3], b * wordbank.final[i][1]]

    print(len(assignment))

    score = {}
    for i in assignment:
        score[i] = (assignment[i][0]+assignment[i][1])

    return score


#def categorization(N, lofw, mu_l, mu_f):


#a = Assembly(example2)
#print(a.max_f)
#print(a.avg_f)
#print(a.dev_f)
#print(a.max_l)
#print(a.avg_l)
#print(a.dev_l)
#print(a.ranges)



example = ["hello.txt"]

example2 = ["harry_1.txt", "harry_2.txt", "harry_3.txt"]

set_of_books = ["harry_1.txt", "harry_2.txt", "harry_3.txt", "harry_4.txt", "harry_5.txt", "harry_6.txt", "harry_7.txt",
                "alice_in_wonderland.txt", "sherlock_holmes.txt", "gullivers_travels.txt", "lotf.txt",
                "mysterious_affair.txt", "oliver_twist.txt", "peter_pan.txt", "scarlet_letter.txt",
                "holes.txt", "tkm.txt", "infinite_jest.txt"]

finite_words = parsed("a_good_amount.txt")
test = Assembly(set_of_books)

from openpyxl import Workbook
wb = Workbook()

ws = wb.active

ws['A1'] = "WORD"
ws['B1'] = "SCORE"

X = comparev2(finite_words, test)
N = len(X)+1
print(X)

for i in X:
    ws.append([i,X[i]])

ws.append(["avg", "=AVERAGE(B2:B{0})".format(N)])

wb.save("word_scores.xlsx")


