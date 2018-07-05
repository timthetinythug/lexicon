from student import Student
from openpyxl import Workbook


# Room will be analogous to a classroom;
# Room has students
# students = {student1: student1_object, student2: student2_object, ...}

class Room:

    def __init__(self, students):
        self.students = students

    def __eq__(self, other):
        for i in self.students:
            if i not in other.students:
                return False
        return True

    def __repr__(self):
        temp = ""
        for i in self.students:
            temp += '\n'+'i'
        return str('attendance' + temp)

    def create_excel_sheets(self):

        for i in self.students:
            student = self.students[i]
            wb = Workbook()
            ws = wb.active

            ws['A1'] = "DATE"
            ws['B1'] = "WORDS"
            ws['C1'] = "NUMBER OF TIMES INCORRECT"

            for k in student.excel: # k = ["some_date", "some_word, 0 or 1", "another_word, 0 or 1", ...]
                date = k[0]

                for word in k[1:]: # word = "some_word, 0 or 1"
                    x = word.split(", ") # x = ["some_word", "0 or 1"]
                    w = x[0]
                    freq = student.vocab[w]
                    ws.append([date, w, freq])

            wb.save("{0}_vocab.xlsx".format(student.name))


